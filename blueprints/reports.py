# -*- coding: utf-8 -*-
"""
보고서 산출물 생성
-------------------
착수보고서 "보안 점수 산출 & 보고서 산출물" 규격을 따른다.

Excel (5시트): 표지 · 대시보드요약 · 항목별상세(위험도순+조건부서식) · 조치이력 · 승인이력
PDF        : 표지 + 요약 + 전체 진단 결과(판정 사유 포함) + 수동확인 항목 + 자동 조치 가능 항목 + 수동 조치 필요 항목
색상 규칙   : 양호=초록 / 취약=빨강 / 수동확인=노랑  (Excel·PDF 동일 적용)
생성 방식   : 대시보드 버튼 클릭 -> 즉시 생성·다운로드 (서버 단위 / 전체 통합 선택 가능)

[수정 이력]
- STATUS 값이 시스템 전반에서 "양호"/"취약"/"수동확인" (한글)을 쓰는데, 보고서 코드에서만
  "N/A"라는 다른 문자열과 비교하고 있어 수동확인 항목이 필터링·색상·안내 문구 어디에도
  반영되지 않던 버그를 수정함 (_pending_manual_count()만 원래 "수동확인"으로 정확히 비교하고
  있었고, 나머지 보고서 본문 코드가 "N/A"로 비교하고 있었음).
- PDF에서 취약 항목이 자동/수동 조치 구분 없이 "상세" 섹션 하나에 섞여 있던 것을,
  "자동 조치 가능 항목" / "수동 조치 필요 항목" 두 섹션으로 명확히 분리함.
- 가이드 텍스트에 있던 "(자동 조치)/(수동 조치)" 라벨은 엑셀의 "조치방식" 컬럼과
  중복이라 제거함 (섹션 제목/컬럼으로 이미 구분되므로).
- PDF가 요약 표 이후 바로 "수동확인 항목"으로 넘어가 취약 항목 전체를 한눈에 보여주는
  개요가 없었음. 엑셀 "항목별상세" 시트를 요약한 "2. 취약 항목" 섹션(서버/코드/카테고리/
  항목명/위험도/조치방식, 위험도순)을 요약과 수동확인 항목 사이에 추가하고 이후 섹션
  번호를 한 칸씩 밀었음.
- "2. 취약 항목"이 취약 건만 보여줘서 양호 항목이 왜 양호인지는 알 수 없었음. 양호/취약/
  수동확인 전체 결과를 보여주는 "2. 전체 진단 결과"로 확장하고, evidence(진단 스크립트의
  판정 근거)를 "판정 사유" 컬럼으로 추가함 (evidence가 없는 경우에만 결과별 기본 문구로
  대체 - _judge_reason()). 조치방식 컬럼은 섹션 4/5에서 이미 다루므로 제거하고 그 자리에
  진단결과/판정 사유 컬럼을 넣음.
"""
import io
import os
from datetime import datetime
from xml.sax.saxutils import escape as _xml_escape

from flask import Blueprint, request, send_file, render_template, flash, redirect, url_for
from flask_login import login_required

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from models import Server, ActionHistory, ApprovalRequest
from scoring import calc_security_level, grade_of, GRADE_COLOR
from impact import get_impact

reports_bp = Blueprint("reports", __name__, url_prefix="/reports")

# 한글 PDF 폰트 등록 (ReportLab 기본 폰트는 한글 글리프가 없어 표지·표가 깨져 나온다)
_FONT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "static", "fonts")
KO_FONT = "NotoSansKR"
KO_FONT_BOLD = "NotoSansKR-Bold"
pdfmetrics.registerFont(TTFont(KO_FONT, os.path.join(_FONT_DIR, "NotoSansKR-Regular.ttf")))
pdfmetrics.registerFont(TTFont(KO_FONT_BOLD, os.path.join(_FONT_DIR, "NotoSansKR-Bold.ttf")))

GOOD_FILL = PatternFill("solid", fgColor="C6EFCE")
BAD_FILL = PatternFill("solid", fgColor="FFC7CE")
NA_FILL = PatternFill("solid", fgColor="FFEB9C")
HEAD_FILL = PatternFill("solid", fgColor="1F2937")
HEAD_FONT = Font(name="Arial", bold=True, color="FFFFFF")
BASE_FONT = Font(name="Arial", size=10)
TITLE_FONT = Font(name="Arial", size=20, bold=True)
SUB_FONT = Font(name="Arial", size=11, color="555555")
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

RISK_ORDER = {"H": 0, "M": 1, "L": 2}
# PDF 표에서 위험도(H/M/L)를 표시할 때 쓰는 배경색 - 엑셀의 양호/취약/수동확인 색상 규칙과
# 동일한 팔레트(빨강 FFC7CE · 노랑 FFEB9C · 초록 C6EFCE)를 재사용해 위험도 순으로 매핑한다.
RISK_FILL_PDF = {
    "H": colors.HexColor("#FFC7CE"),
    "M": colors.HexColor("#FFEB9C"),
    "L": colors.HexColor("#C6EFCE"),
}

# 진단 결과 상태값 상수 (scan.yaml / 각 check.sh의 STATUS와 동일한 문자열이어야 함)
STATUS_GOOD = "양호"
STATUS_BAD = "취약"
STATUS_MANUAL_CHECK = "수동확인"

# PDF 표에서 진단결과(양호/취약/수동확인)를 표시할 때 쓰는 배경색 - 엑셀 _result_fill()과
# 동일한 팔레트(초록/빨강/노랑)를 재사용한다.
RESULT_FILL_PDF = {
    STATUS_GOOD: colors.HexColor("#C6EFCE"),
    STATUS_BAD: colors.HexColor("#FFC7CE"),
    STATUS_MANUAL_CHECK: colors.HexColor("#FFEB9C"),
}


def _target_servers():
    ids = request.args.getlist("server_id")
    if not ids or "all" in ids:
        return Server.query.order_by(Server.id).all(), "전체 서버"
    if len(ids) == 1:
        srv = Server.query.get_or_404(int(ids[0]))
        return [srv], srv.hostname
    servers = Server.query.filter(Server.id.in_([int(i) for i in ids])).order_by(Server.id).all()
    label = ", ".join(s.hostname for s in servers) if len(servers) <= 4 else f"선택 서버 {len(servers)}대"
    return servers, label


def _style_header(ws, row, headers, widths):
    for idx, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=idx, value=h)
        c.font = HEAD_FONT
        c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
        ws.column_dimensions[get_column_letter(idx)].width = widths[idx - 1]


def _result_fill(result):
    if result == STATUS_BAD:
        return BAD_FILL
    if result == STATUS_MANUAL_CHECK:
        return NA_FILL
    return GOOD_FILL


def _impact_text(check_code):
    imp = get_impact(check_code)
    return f"[{imp['level_label']}] {imp['apply']} · {imp['scope']}"


def _judge_reason(r):
    """양호/취약/수동확인 판정 근거 텍스트. evidence가 진단 스크립트의 판정 근거를
    그대로 담고 있으므로 우선 사용하고, 없을 때만 결과별 기본 문구로 대체한다."""
    if r.evidence:
        return r.evidence
    if r.result == STATUS_BAD:
        return "진단 스크립트가 점검 기준을 충족하지 못한 것으로 판정했습니다."
    if r.result == STATUS_MANUAL_CHECK:
        return "자동 진단 스크립트가 양호/취약을 확정 판정하지 못해 관리자의 수동 확인이 필요합니다."
    return "진단 스크립트가 점검 기준을 충족하는 것으로 판정했습니다."


def _guide_and_impact(r):
    """취약(자동/수동) 항목의 조치 가이드와 조치 시 영향도 텍스트.
    자동/수동 구분은 별도 컬럼(조치방식)이나 섹션 제목으로 이미 표시되므로
    가이드 텍스트 자체에는 라벨을 넣지 않는다."""
    ci = r.check_item
    guide_text = ci.guide or r.recommendation or "-"
    return guide_text, _impact_text(ci.code)


def _na_guidance(r):
    """수동 확인(수동확인) 항목의 확인 사유 / 확인 방법 / 조치 방법 / 조치 시 영향도 텍스트."""
    ci = r.check_item
    reason = r.evidence or "자동 진단 스크립트가 양호/취약을 확정 판정하지 못해 관리자의 수동 확인이 필요합니다."
    how = (f"다음 기준과 비교해 현재 설정을 직접 확인: {r.recommendation}"
           if r.recommendation else "가이드 기준에 따라 현재 설정을 직접 확인합니다.")
    fix = ci.guide or r.recommendation or "-"
    return reason, how, fix, _impact_text(ci.code)


def _diagnosis_narrative(servers):
    """진단 결과 총평을 규칙 기반으로 문단 몇 개짜리 줄글로 만든다. AI/LLM 호출 없이
    이미 계산되는 통계(평균점수·카테고리별 준수율·위험도별 건수 등)를 문장 틀에 끼워
    넣는 방식이라 결정적이고(같은 데이터면 항상 같은 문장) 빠르다."""
    if not servers:
        return ["진단 대상 서버가 없습니다."]

    per_server = []
    for srv in servers:
        results = srv.latest_results()
        level, grade, total, identified = calc_security_level(results)
        per_server.append((srv, level, grade, results))
    avg = round(sum(p[1] for p in per_server) / len(per_server), 1)
    worst = min(per_server, key=lambda p: p[1])
    best = max(per_server, key=lambda p: p[1])

    all_pairs = [(srv, r) for srv, _, _, results in per_server for r in results]
    vuln_pairs = [(s, r) for s, r in all_pairs if r.result == STATUS_BAD]
    manual_pairs = [(s, r) for s, r in all_pairs if r.result == STATUS_MANUAL_CHECK]
    high_vuln = [x for x in vuln_pairs if x[1].check_item.risk_level == "H"]
    auto_fixable = [x for x in vuln_pairs if x[1].remediation_type == "auto"]
    manual_fixable = [x for x in vuln_pairs if x[1].remediation_type == "manual"]

    cat_stats = {}
    for s, r in all_pairs:
        if r.result == STATUS_MANUAL_CHECK:
            continue
        cat_stats.setdefault(r.check_item.category, [0, 0])
        cat_stats[r.check_item.category][0] += 1
        if r.result == STATUS_BAD:
            cat_stats[r.check_item.category][1] += 1
    worst_cat, worst_rate = None, 101
    for cat, (tot, vuln) in cat_stats.items():
        if tot == 0:
            continue
        rate = (1 - vuln / tot) * 100
        if rate < worst_rate:
            worst_cat, worst_rate = cat, rate

    paras = [
        f"이번 진단은 총 {len(servers)}대 서버를 대상으로 진행되었으며, 평균 보안점수는 "
        f"{avg}점({grade_of(avg)} 등급)입니다. 보안 수준이 가장 낮은 서버는 "
        f"{worst[0].hostname}({worst[1]}점, {worst[2]} 등급)이며, 가장 높은 서버는 "
        f"{best[0].hostname}({best[1]}점, {best[2]} 등급)입니다."
    ]
    if vuln_pairs:
        paras.append(
            f"전체 {len(vuln_pairs)}건의 취약 항목이 확인되었고, 이 중 위험도 '상' 항목이 "
            f"{len(high_vuln)}건으로 가장 시급한 조치 대상입니다. 조치 방식 기준으로는 자동 조치가 "
            f"가능한 항목이 {len(auto_fixable)}건, 실무자가 직접 조치해야 하는 항목이 "
            f"{len(manual_fixable)}건입니다."
        )
    else:
        paras.append("확인된 취약 항목이 없어 전 항목이 양호한 상태입니다.")
    if worst_cat:
        paras.append(
            f"카테고리별로는 '{worst_cat}' 영역의 준수율이 {round(worst_rate, 1)}%로 가장 낮게 나타나, "
            f"해당 영역을 우선 점검·조치하는 것을 권장합니다."
        )
    if manual_pairs:
        paras.append(
            f"이 외에 자동 진단 스크립트가 양호/취약을 확정하지 못해 관리자의 확인이 필요한 "
            f"수동확인 항목이 {len(manual_pairs)}건 남아 있습니다."
        )
    return paras


def _action_narrative(hist_rows):
    """조치 보고서(완료된 조치만)용 총평 문단. 마찬가지로 규칙 기반, AI 호출 없음."""
    if not hist_rows:
        return ["이번 범위에서 완료된 조치 이력이 없습니다."]
    auto_cnt = sum(1 for h in hist_rows if (h.approval.remediation_type if h.approval else
                                             ("manual" if h.work_note else "auto")) == "auto")
    manual_cnt = len(hist_rows) - auto_cnt
    score_delta = sum((h.before_score or 0) - (h.after_score or 0) for h in hist_rows)
    servers_touched = len({h.server_id for h in hist_rows})
    paras = [
        f"이번 범위에서 총 {len(hist_rows)}건의 조치가 완료되었습니다 (자동 조치 {auto_cnt}건, "
        f"수동 조치 {manual_cnt}건), 대상 서버는 {servers_touched}대입니다."
    ]
    if score_delta > 0:
        paras.append(
            f"조치 전후 배점 기준으로 총 {round(score_delta, 1)}점 상당의 취약점 배점이 해소되어, "
            f"조치 대상 서버들의 보안 수준이 개선되었습니다."
        )
    return paras


def _pending_manual_count():
    """자동 진단 스크립트가 판정 못해 사람이 확인해야 하는(수동확인) 항목이 아직 남았는지.
    이게 남아있으면 진단 결과가 아직 확정이 아니라서 진단 보고서를 뽑을 수 없다."""
    count = 0
    for srv in Server.query.all():
        for r in srv.latest_results():
            if r.result == STATUS_MANUAL_CHECK:
                count += 1
    return count


def _pending_action_count():
    """승인은 됐지만 아직 실행되지 않은 조치 건수 - 자동 조치는 관리자의 '자동 조치 실행'
    대기중, 수동 조치는 실무자의 '완료 처리' 대기중인 것 모두 포함한다. 이게 남아있으면
    조치가 실제로는 아직 안 끝난 거라 조치 보고서를 뽑을 수 없다."""
    return (ApprovalRequest.query.filter_by(status="approved")
            .filter(ApprovalRequest.executed_at.is_(None)).count())


@reports_bp.route("/")
@login_required
def index():
    servers = Server.query.order_by(Server.id).all()
    return render_template(
        "reports.html", servers=servers,
        pending_manual=_pending_manual_count(),
        pending_action=_pending_action_count(),
    )


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------
@reports_bp.route("/excel")
@login_required
def excel_report():
    pending = _pending_manual_count()
    if pending:
        flash(f"수동확인 대기 항목이 {pending}건 남아있어 진단 보고서를 생성할 수 없습니다. "
              f"먼저 수동확인을 모두 판정하세요.", "warning")
        return redirect(url_for("reports.index"))

    servers, scope_label = _target_servers()
    wb = Workbook()

    # 1) 표지 -----------------------------------------------------------
    cover = wb.active
    cover.title = "표지"
    cover.sheet_view.showGridLines = False
    cover["B3"] = "KISA U-01~U-67 다중 서버 보안 자동화 플랫폼"
    cover["B3"].font = Font(name="Arial", size=13, color="888888")
    cover["B5"] = "서버 취약점 진단 결과 보고서"
    cover["B5"].font = TITLE_FONT
    cover["B7"] = f"대상 범위 : {scope_label}"
    cover["B7"].font = SUB_FONT
    cover["B8"] = f"생성 일시 : {datetime.now():%Y-%m-%d %H:%M}"
    cover["B8"].font = SUB_FONT
    cover["B9"] = "생성 방식 : 대시보드 버튼 클릭 즉시 생성"
    cover["B9"].font = SUB_FONT
    cover["B11"] = "색상 규칙 : 양호=초록 · 취약=빨강 · 수동확인=노랑"
    cover["B11"].font = SUB_FONT
    for col, w in zip("ABCDEFG", [4, 60, 16, 16, 16, 16, 16]):
        cover.column_dimensions[col].width = w

    # 2) 대시보드요약 ------------------------------------------------------
    ws = wb.create_sheet("대시보드요약")
    ws["B2"] = "서버별 보안 수준 요약"
    ws["B2"].font = Font(name="Arial", size=13, bold=True)
    headers = ["호스트명", "IP", "OS", "업무명", "보안점수", "등급", "취약 건수", "전체 항목수"]
    widths = [16, 16, 18, 22, 12, 10, 12, 12]
    _style_header(ws, 4, headers, widths)

    row = 5
    total_score = 0
    for srv in servers:
        results = srv.latest_results()
        level, grade, total, identified = calc_security_level(results)
        vuln_cnt = sum(1 for r in results if r.result == STATUS_BAD)
        applicable = sum(1 for r in results if r.result != STATUS_MANUAL_CHECK)
        ws.cell(row=row, column=1, value=srv.hostname).border = BORDER
        ws.cell(row=row, column=2, value=srv.ip).border = BORDER
        ws.cell(row=row, column=3, value=srv.os_label).border = BORDER
        ws.cell(row=row, column=4, value=srv.business_name).border = BORDER
        sc = ws.cell(row=row, column=5, value=level)
        sc.border = BORDER
        sc.fill = GOOD_FILL if grade in ("우수", "양호") else (NA_FILL if grade == "보통" else BAD_FILL)
        ws.cell(row=row, column=6, value=grade).border = BORDER
        ws.cell(row=row, column=7, value=vuln_cnt).border = BORDER
        ws.cell(row=row, column=8, value=applicable).border = BORDER
        total_score += level
        row += 1

    avg_row = row + 1
    ws.cell(row=avg_row, column=4, value="전체 평균").font = Font(name="Arial", bold=True)
    avg_cell = ws.cell(row=avg_row, column=5,
                        value=f"=ROUND(AVERAGE(E5:E{row-1}),1)" if row > 5 else 0)
    avg_cell.font = Font(name="Arial", bold=True)

    # 카테고리별 준수율
    cat_row = avg_row + 3
    ws.cell(row=cat_row, column=2, value="카테고리별 준수율").font = Font(name="Arial", size=12, bold=True)
    cat_headers = ["카테고리", "전체 항목수(서버합산)", "취약 건수", "준수율(%)"]
    _style_header(ws, cat_row + 1, cat_headers, [24, 20, 14, 14])
    cat_stats = {}
    for srv in servers:
        for r in srv.latest_results():
            if r.result == STATUS_MANUAL_CHECK:
                continue
            cat = r.check_item.category
            cat_stats.setdefault(cat, [0, 0])
            cat_stats[cat][0] += 1
            if r.result == STATUS_BAD:
                cat_stats[cat][1] += 1
    r_i = cat_row + 2
    for cat, (tot, vuln) in cat_stats.items():
        rate = round((1 - vuln / tot) * 100, 1) if tot else 0
        ws.cell(row=r_i, column=2, value=cat).border = BORDER
        ws.cell(row=r_i, column=3, value=tot).border = BORDER
        ws.cell(row=r_i, column=4, value=vuln).border = BORDER
        rc = ws.cell(row=r_i, column=5, value=rate)
        rc.border = BORDER
        rc.fill = GOOD_FILL if rate >= 80 else (NA_FILL if rate >= 60 else BAD_FILL)
        r_i += 1

    # 3) 항목별상세 (위험도순 + 조건부서식) ---------------------------------
    #    - 취약(자동조치 가능): 조치 시 무엇이 바뀌는지 + 조치 시 영향도
    #    - 취약(수동조치 필요): 수동 조치 가이드 + 조치 시 영향도
    #    - 수동확인: 확인 사유 + 확인 방법 + 조치 방법 + 조치 시 영향도
    ws2 = wb.create_sheet("항목별상세")
    # PDF "2. 전체 진단 결과"와 항목·컬럼을 동일하게 맞춘다 (판정사유=evidence 포함).
    headers2 = ["서버", "항목코드", "카테고리", "점검항목", "위험도", "진단결과", "판정사유", "배점",
                "현재설정(취약시)", "조치권고", "조치방식", "가이드(조치 내용 / 확인 사유·방법)", "영향도"]
    widths2 = [14, 10, 20, 34, 8, 10, 50, 8, 40, 40, 10, 50, 44]
    _style_header(ws2, 1, headers2, widths2)

    all_rows = []
    for srv in servers:
        for r in srv.latest_results():
            ci = r.check_item
            all_rows.append((srv, r, ci))
    all_rows.sort(key=lambda x: (RISK_ORDER.get(x[2].risk_level, 9), x[0].hostname, x[2].order_no))

    r_i = 2
    for srv, r, ci in all_rows:
        remediation_label = ("수동" if r.remediation_type == "manual" else "자동") if r.result == STATUS_BAD else ""
        if r.result == STATUS_BAD:
            guide_text, impact_text = _guide_and_impact(r)
        elif r.result == STATUS_MANUAL_CHECK:
            reason, how, fix, impact_text = _na_guidance(r)
            guide_text = f"[확인 사유] {reason}\n[확인 방법] {how}\n[조치 방법] {fix}"
        else:
            guide_text, impact_text = "-", "-"
        vals = [srv.hostname, ci.code, ci.category, ci.name, ci.risk_level, r.result, _judge_reason(r),
                ci.weight, r.current_setting or "", r.recommendation or "", remediation_label,
                guide_text, impact_text]
        for col, v in enumerate(vals, start=1):
            cell = ws2.cell(row=r_i, column=col, value=v)
            cell.font = BASE_FONT
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=(col in (7, 9, 10, 12, 13)), vertical="top")
            if col == 6:
                cell.fill = _result_fill(r.result)
        r_i += 1
    ws2.freeze_panes = "A2"

    # 4) 조치이력 --------------------------------------------------------
    ws3 = wb.create_sheet("조치이력")
    headers3 = ["수행일시", "서버", "항목코드", "항목명", "이전결과", "이후결과",
                "이전점수", "이후점수", "수행자", "백업경로"]
    widths3 = [18, 14, 10, 30, 10, 10, 10, 10, 26, 34]
    _style_header(ws3, 1, headers3, widths3)
    server_ids = {s.id for s in servers}
    hist = (ActionHistory.query.order_by(ActionHistory.performed_at.desc()).all())
    r_i = 2
    for h in hist:
        if h.server_id not in server_ids:
            continue
        vals = [h.performed_at.strftime("%Y-%m-%d %H:%M"), h.server.hostname, h.check_item.code,
                h.check_item.name, h.before_result, h.after_result, h.before_score, h.after_score,
                h.performed_by, h.backup_path]
        for col, v in enumerate(vals, start=1):
            cell = ws3.cell(row=r_i, column=col, value=v)
            cell.font = BASE_FONT
            cell.border = BORDER
        r_i += 1

    # 5) 승인이력 --------------------------------------------------------
    ws4 = wb.create_sheet("승인이력")
    headers4 = ["요청일시", "서버", "항목코드", "항목명", "요청자", "상태", "승인자",
                "결정일시", "거부사유", "실행일시"]
    widths4 = [18, 14, 10, 30, 16, 10, 16, 18, 30, 18]
    _style_header(ws4, 1, headers4, widths4)
    apr = ApprovalRequest.query.order_by(ApprovalRequest.requested_at.desc()).all()
    r_i = 2
    status_label = {"pending": "대기", "approved": "승인", "rejected": "거부"}
    for a in apr:
        if a.server_id not in server_ids:
            continue
        vals = [a.requested_at.strftime("%Y-%m-%d %H:%M"), a.server.hostname, a.check_item.code,
                a.check_item.name, a.requested_by, status_label.get(a.status, a.status),
                a.approver or "", a.decided_at.strftime("%Y-%m-%d %H:%M") if a.decided_at else "",
                a.reject_reason or "", a.executed_at.strftime("%Y-%m-%d %H:%M") if a.executed_at else ""]
        for col, v in enumerate(vals, start=1):
            cell = ws4.cell(row=r_i, column=col, value=v)
            cell.font = BASE_FONT
            cell.border = BORDER
            if col == 6:
                cell.fill = {"대기": NA_FILL, "승인": GOOD_FILL, "거부": BAD_FILL}.get(v, PatternFill())
        r_i += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"보안진단_보고서_{scope_label}_{datetime.now():%Y%m%d_%H%M}.xlsx".replace(" ", "")
    return send_file(buf, as_attachment=True, download_name=fname,
                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ---------------------------------------------------------------------------
# PDF (markdown -> HTML -> weasyprint 대신, 배포 편의를 위해 reportlab으로 직접 생성)
# ---------------------------------------------------------------------------
@reports_bp.route("/pdf")
@login_required
def pdf_report():
    pending = _pending_manual_count()
    if pending:
        flash(f"수동확인 대기 항목이 {pending}건 남아있어 진단 보고서를 생성할 수 없습니다. "
              f"먼저 수동확인을 모두 판정하세요.", "warning")
        return redirect(url_for("reports.index"))

    servers, scope_label = _target_servers()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                             topMargin=18 * mm, bottomMargin=16 * mm,
                             leftMargin=16 * mm, rightMargin=16 * mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleKo", parent=styles["Title"], fontName=KO_FONT_BOLD, fontSize=20, leading=24)
    h2_style = ParagraphStyle("H2Ko", parent=styles["Heading2"], fontName=KO_FONT_BOLD, fontSize=13, spaceBefore=14, spaceAfter=6)
    normal = ParagraphStyle("NormalKo", parent=styles["Normal"], fontName=KO_FONT)
    cell_style = ParagraphStyle("CellKo", parent=normal, fontSize=7, leading=9)
    narrative_style = ParagraphStyle("NarrativeKo", parent=normal, fontSize=10, leading=15, spaceAfter=7)

    # 임베드한 NotoSansKR 폰트에 없는 글리프(가운뎃점 · 등)는 PDF에서 빈 네모(tofu)로
    # 깨져 나오므로, 폰트가 지원하는 문자로 치환해 셀에 넣는다.
    _PDF_UNSUPPORTED_GLYPHS = {"·": " / ", "→": "->", "×": "x"}

    def _sub_glyphs(text):
        t = (text or "").strip() or "-"
        for ch, repl in _PDF_UNSUPPORTED_GLYPHS.items():
            t = t.replace(ch, repl)
        return t

    def _cell(text):
        return Paragraph(_xml_escape(_sub_glyphs(text)), cell_style)

    def _narr(text):
        return Paragraph(_xml_escape(_sub_glyphs(text)), narrative_style)

    story = []
    story.append(Paragraph("서버 취약점 진단 결과 보고서", title_style))
    story.append(Paragraph("KISA U-01~U-67 다중 서버 보안 자동화 플랫폼", normal))
    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph(f"대상 범위: {scope_label}", normal))
    story.append(Paragraph(f"생성 일시: {datetime.now():%Y-%m-%d %H:%M}", normal))
    story.append(Spacer(1, 6 * mm))

    # 1. 요약 표
    story.append(Paragraph("1. 요약", h2_style))
    data = [["호스트명", "OS", "보안점수", "등급", "취약 건수"]]
    for srv in servers:
        results = srv.latest_results()
        level, grade, total, identified = calc_security_level(results)
        vuln_cnt = sum(1 for r in results if r.result == STATUS_BAD)
        data.append([srv.hostname, srv.os_label, str(level), grade, str(vuln_cnt)])
    t = Table(data, colWidths=[32 * mm, 30 * mm, 26 * mm, 24 * mm, 26 * mm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), KO_FONT_BOLD),
        ("FONTNAME", (0, 1), (-1, -1), KO_FONT),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D0D0D0")),
        ("ALIGN", (2, 1), (-1, -1), "CENTER"),
    ]))
    story.append(t)
    story.append(Spacer(1, 4 * mm))

    # 총평 - 위 요약 표의 숫자를 문장으로 풀어쓴 줄글. AI 호출 없이 규칙 기반으로
    # 생성한다 (_diagnosis_narrative 참고).
    story.append(Paragraph("총평", h2_style))
    for p in _diagnosis_narrative(servers):
        story.append(_narr(p))
    story.append(Spacer(1, 4 * mm))

    # 2. 전체 진단 결과 - 엑셀 "항목별상세" 시트를 위험도순으로 요약한 목록.
    #    양호/취약/수동확인 전체를 다루며, 각 항목이 왜 그렇게 판정됐는지(판정 사유)도 포함한다.
    story.append(Paragraph("2. 전체 진단 결과", h2_style))
    all_result_rows = []
    for srv in servers:
        for r in srv.latest_results():
            all_result_rows.append((srv, r))
    all_result_rows.sort(key=lambda x: (RISK_ORDER.get(x[1].check_item.risk_level, 9), x[0].hostname,
                                         x[1].check_item.order_no))
    all_data = [["서버", "코드", "카테고리", "항목명", "위험도", "진단결과", "판정 사유"]]
    for srv, r in all_result_rows:
        ci = r.check_item
        all_data.append([_cell(srv.hostname), _cell(ci.code), _cell(ci.category), _cell(ci.name),
                          ci.risk_level, r.result, _cell(_judge_reason(r))])
    if len(all_data) == 1:
        all_data.append(["-", "-", "-", "진단 결과 없음", "-", "-", "-"])
    t_all = Table(all_data, colWidths=[24 * mm, 12 * mm, 22 * mm, 38 * mm, 12 * mm, 16 * mm, 54 * mm])
    style_all = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), KO_FONT_BOLD),
        ("FONTNAME", (0, 1), (-1, -1), KO_FONT),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D0D0D0")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (4, 1), (5, -1), "CENTER"),
    ]
    for i, (srv, r) in enumerate(all_result_rows, start=1):
        risk_fill = RISK_FILL_PDF.get(r.check_item.risk_level)
        if risk_fill:
            style_all.append(("BACKGROUND", (4, i), (4, i), risk_fill))
        result_fill = RESULT_FILL_PDF.get(r.result)
        if result_fill:
            style_all.append(("BACKGROUND", (5, i), (5, i), result_fill))
    t_all.setStyle(TableStyle(style_all))
    story.append(t_all)
    story.append(Spacer(1, 6 * mm))

    # 3. 수동 확인 필요 항목 - 왜 필요한지 / 어떻게 확인하는지 / 조치 방법 / 조치 시 영향도
    story.append(Paragraph("3. 수동 확인 필요 항목", h2_style))
    na_rows = []
    for srv in servers:
        for r in srv.latest_results():
            if r.result == STATUS_MANUAL_CHECK:
                na_rows.append((srv, r))
    if na_rows:
        na_data = [["서버", "코드", "항목명", "확인 사유", "확인 방법", "조치 방법", "영향도"]]
        for srv, r in na_rows:
            reason, how, fix, impact_text = _na_guidance(r)
            na_data.append([_cell(srv.hostname), _cell(r.check_item.code), _cell(r.check_item.name),
                             _cell(reason), _cell(how), _cell(fix), _cell(impact_text)])
        t3 = Table(na_data, colWidths=[26 * mm, 10 * mm, 18 * mm, 28 * mm, 28 * mm, 26 * mm, 38 * mm])
        t3.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), KO_FONT_BOLD),
            ("FONTNAME", (0, 1), (-1, -1), KO_FONT),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D0D0D0")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("BACKGROUND", (0, 1), (2, -1), colors.HexColor("#FFEB9C")),
        ]))
        story.append(t3)
    else:
        story.append(Paragraph("수동 확인이 필요한 항목이 없습니다.", normal))
    story.append(Spacer(1, 6 * mm))

    # 4. 자동 조치 가능 항목 - 조치 시 무엇이 바뀌는지 / 조치 시 영향도
    story.append(Paragraph("4. 자동 조치 가능 항목", h2_style))
    auto_data = [["서버", "코드", "항목명", "위험도", "조치 내용(적용 시 변경 사항)", "조치 시 영향도"]]
    auto_rows = []
    for srv in servers:
        for r in srv.latest_results():
            if r.result == STATUS_BAD and r.remediation_type == "auto":
                auto_rows.append((srv, r))
    auto_rows.sort(key=lambda x: (RISK_ORDER.get(x[1].check_item.risk_level, 9), x[0].hostname))
    for srv, r in auto_rows:
        guide_text, impact_text = _guide_and_impact(r)
        auto_data.append([_cell(srv.hostname), _cell(r.check_item.code), _cell(r.check_item.name),
                           r.check_item.risk_level, _cell(guide_text), _cell(impact_text)])
    if len(auto_data) == 1:
        auto_data.append(["-", "-", "자동 조치 가능 항목 없음", "-", "-", "-"])
    t_auto = Table(auto_data, colWidths=[26 * mm, 12 * mm, 26 * mm, 10 * mm, 52 * mm, 52 * mm])
    style_auto = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), KO_FONT_BOLD),
        ("FONTNAME", (0, 1), (-1, -1), KO_FONT),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D0D0D0")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]
    for i, (srv, r) in enumerate(auto_rows, start=1):
        fill = RISK_FILL_PDF.get(r.check_item.risk_level)
        if fill:
            style_auto.append(("BACKGROUND", (3, i), (3, i), fill))
    t_auto.setStyle(TableStyle(style_auto))
    story.append(t_auto)
    story.append(Spacer(1, 6 * mm))

    # 5. 수동 조치 필요 항목 - 가이드 / 영향도
    story.append(Paragraph("5. 수동 조치 필요 항목", h2_style))
    manual_rows = []
    for srv in servers:
        for r in srv.latest_results():
            if r.result == STATUS_BAD and r.remediation_type == "manual":
                manual_rows.append((srv, r))
    manual_rows.sort(key=lambda x: (RISK_ORDER.get(x[1].check_item.risk_level, 9), x[0].hostname))
    manual_data = [["서버", "코드", "항목명", "위험도", "수동 조치 가이드", "영향도"]]
    for srv, r in manual_rows:
        guide_text, impact_text = _guide_and_impact(r)
        manual_data.append([_cell(srv.hostname), _cell(r.check_item.code), _cell(r.check_item.name),
                             r.check_item.risk_level, _cell(guide_text), _cell(impact_text)])
    if len(manual_data) == 1:
        manual_data.append(["-", "-", "수동 조치가 필요한 항목 없음", "-", "-", "-"])
    t_manual = Table(manual_data, colWidths=[30 * mm, 12 * mm, 26 * mm, 14 * mm, 52 * mm, 44 * mm])
    style_manual = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), KO_FONT_BOLD),
        ("FONTNAME", (0, 1), (-1, -1), KO_FONT),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D0D0D0")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]
    for i, (srv, r) in enumerate(manual_rows, start=1):
        fill = RISK_FILL_PDF.get(r.check_item.risk_level)
        if fill:
            style_manual.append(("BACKGROUND", (3, i), (3, i), fill))
    t_manual.setStyle(TableStyle(style_manual))
    story.append(t_manual)

    doc.build(story)
    buf.seek(0)
    fname = f"보안진단_보고서_{scope_label}_{datetime.now():%Y%m%d_%H%M}.pdf".replace(" ", "")
    return send_file(buf, as_attachment=True, download_name=fname, mimetype="application/pdf")


# ---------------------------------------------------------------------------
# 조치 보고서 (Excel) - 진단 결과가 아니라 "실제로 완료된 조치"만 다룬다.
# ActionHistory는 자동 조치(승인 즉시 실행) 또는 수동 조치(실무자가 조치 내용을 남기고
# '완료 처리'한 건)에서만 생성되므로, 이 테이블을 그대로 쓰면 완료되지 않은 건은
# 자연히 빠진다 - 별도 필터가 필요 없다.
# ---------------------------------------------------------------------------
@reports_bp.route("/action/excel")
@login_required
def action_excel_report():
    pending = _pending_action_count()
    if pending:
        flash(f"실행 대기 중인 조치가 {pending}건 있어 조치 보고서를 생성할 수 없습니다 "
              f"(자동 조치는 관리자의 실행, 수동 조치는 실무자의 완료 처리를 먼저 마쳐야 합니다).", "warning")
        return redirect(url_for("reports.index"))

    servers, scope_label = _target_servers()
    server_ids = {s.id for s in servers}

    wb = Workbook()
    ws = wb.active
    ws.title = "조치 보고서"
    ws["B2"] = "조치 이행 결과 보고서"
    ws["B2"].font = TITLE_FONT
    ws["B3"] = f"대상 범위 : {scope_label}"
    ws["B3"].font = SUB_FONT
    ws["B4"] = f"생성 일시 : {datetime.now():%Y-%m-%d %H:%M}"
    ws["B4"].font = SUB_FONT
    ws["B5"] = "완료된 조치만 포함됩니다 (자동=승인 즉시 실행 / 수동=실무자가 조치 내용을 남기고 완료 처리)"
    ws["B5"].font = SUB_FONT

    headers = ["수행일시", "서버", "항목코드", "항목명", "위험도", "조치방식",
               "이전결과", "이후결과", "이전점수", "이후점수", "수행자", "조치 내용", "백업경로"]
    widths = [18, 14, 10, 30, 8, 10, 10, 10, 10, 10, 26, 46, 34]
    header_row = 7
    _style_header(ws, header_row, headers, widths)

    hist = ActionHistory.query.order_by(ActionHistory.performed_at.desc()).all()
    r_i = header_row + 1
    for h in hist:
        if h.server_id not in server_ids:
            continue
        remediation_type = h.approval.remediation_type if h.approval else ("manual" if h.work_note else "auto")
        remediation_label = "수동" if remediation_type == "manual" else "자동"
        vals = [h.performed_at.strftime("%Y-%m-%d %H:%M"), h.server.hostname, h.check_item.code,
                h.check_item.name, h.check_item.risk_level, remediation_label,
                h.before_result, h.after_result, h.before_score, h.after_score,
                h.performed_by, h.work_note or "-", h.backup_path]
        for col, v in enumerate(vals, start=1):
            cell = ws.cell(row=r_i, column=col, value=v)
            cell.font = BASE_FONT
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=(col == 12), vertical="top")
        r_i += 1
    ws.freeze_panes = f"A{header_row + 1}"

    if r_i == header_row + 1:
        ws.cell(row=r_i, column=2, value="완료된 조치 이력이 없습니다.").font = SUB_FONT

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"조치_보고서_{scope_label}_{datetime.now():%Y%m%d_%H%M}.xlsx".replace(" ", "")
    return send_file(buf, as_attachment=True, download_name=fname,
                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ---------------------------------------------------------------------------
# 조치 보고서 (PDF) - 위 Excel과 항목·컬럼을 동일하게 맞춘 PDF 버전.
# 컬럼이 13개라 A4 세로에는 안 들어가서 가로(landscape)로 만든다.
# ---------------------------------------------------------------------------
@reports_bp.route("/action/pdf")
@login_required
def action_pdf_report():
    pending = _pending_action_count()
    if pending:
        flash(f"실행 대기 중인 조치가 {pending}건 있어 조치 보고서를 생성할 수 없습니다 "
              f"(자동 조치는 관리자의 실행, 수동 조치는 실무자의 완료 처리를 먼저 마쳐야 합니다).", "warning")
        return redirect(url_for("reports.index"))

    servers, scope_label = _target_servers()
    server_ids = {s.id for s in servers}

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                             topMargin=14 * mm, bottomMargin=14 * mm,
                             leftMargin=14 * mm, rightMargin=14 * mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleKo2", parent=styles["Title"], fontName=KO_FONT_BOLD, fontSize=20, leading=24)
    h2_style = ParagraphStyle("H2Ko2", parent=styles["Heading2"], fontName=KO_FONT_BOLD, fontSize=13, spaceBefore=10, spaceAfter=6)
    normal = ParagraphStyle("NormalKo2", parent=styles["Normal"], fontName=KO_FONT)
    cell_style = ParagraphStyle("CellKo2", parent=normal, fontSize=7, leading=9)
    narrative_style = ParagraphStyle("NarrativeKo2", parent=normal, fontSize=10, leading=15, spaceAfter=7)

    _PDF_UNSUPPORTED_GLYPHS = {"·": " / ", "→": "->", "×": "x"}

    def _sub_glyphs(text):
        t = (text or "").strip() or "-"
        for ch, repl in _PDF_UNSUPPORTED_GLYPHS.items():
            t = t.replace(ch, repl)
        return t

    def _cell(text):
        return Paragraph(_xml_escape(_sub_glyphs(text)), cell_style)

    def _narr(text):
        return Paragraph(_xml_escape(_sub_glyphs(text)), narrative_style)

    story = []
    story.append(Paragraph("조치 이행 결과 보고서", title_style))
    story.append(Paragraph("KISA U-01~U-67 다중 서버 보안 자동화 플랫폼", normal))
    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph(f"대상 범위: {scope_label}", normal))
    story.append(Paragraph(f"생성 일시: {datetime.now():%Y-%m-%d %H:%M}", normal))
    story.append(Paragraph("완료된 조치만 포함됩니다 (자동=관리자 실행 완료 / 수동=실무자 조치 내용 기록 후 완료 처리)", normal))
    story.append(Spacer(1, 5 * mm))

    hist = [h for h in ActionHistory.query.order_by(ActionHistory.performed_at.desc()).all()
            if h.server_id in server_ids]

    story.append(Paragraph("총평", h2_style))
    for p in _action_narrative(hist):
        story.append(_narr(p))
    story.append(Spacer(1, 4 * mm))

    story.append(Paragraph("조치 이행 목록", h2_style))
    headers = ["수행일시", "서버", "항목코드", "항목명", "위험도", "조치방식",
               "이전결과", "이후결과", "이전점수", "이후점수", "수행자", "조치 내용", "백업경로"]
    data = [headers]
    for h in hist:
        remediation_type = h.approval.remediation_type if h.approval else ("manual" if h.work_note else "auto")
        remediation_label = "수동" if remediation_type == "manual" else "자동"
        data.append([
            _cell(h.performed_at.strftime("%Y-%m-%d %H:%M")), _cell(h.server.hostname),
            _cell(h.check_item.code), _cell(h.check_item.name), h.check_item.risk_level,
            remediation_label, h.before_result, h.after_result,
            str(h.before_score), str(h.after_score), _cell(h.performed_by),
            _cell(h.work_note or "-"), _cell(h.backup_path),
        ])
    if len(data) == 1:
        data.append(["-", "-", "-", "완료된 조치 이력이 없습니다.", "-", "-", "-", "-", "-", "-", "-", "-", "-"])
    col_widths = [22 * mm, 18 * mm, 14 * mm, 30 * mm, 10 * mm, 12 * mm, 14 * mm, 14 * mm,
                  12 * mm, 12 * mm, 22 * mm, 40 * mm, 40 * mm]
    t = Table(data, colWidths=col_widths, repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), KO_FONT_BOLD),
        ("FONTNAME", (0, 1), (-1, -1), KO_FONT),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D0D0D0")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (4, 1), (9, -1), "CENTER"),
    ]
    for i in range(1, len(data)):
        fill = RISK_FILL_PDF.get(data[i][4])
        if fill:
            style.append(("BACKGROUND", (4, i), (4, i), fill))
    t.setStyle(TableStyle(style))
    story.append(t)

    doc.build(story)
    buf.seek(0)
    fname = f"조치_보고서_{scope_label}_{datetime.now():%Y%m%d_%H%M}.pdf".replace(" ", "")
    return send_file(buf, as_attachment=True, download_name=fname, mimetype="application/pdf")